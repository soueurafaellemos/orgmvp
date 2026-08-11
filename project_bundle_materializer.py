from __future__ import annotations

import csv
import io
import json
import math
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


from document_io import prepare_documents as prepare_input_documents
from memory_briefing import analyze_briefing_document
from memory_cost_parser import parse_cost_workbook
from memory_db import save_memory_presentation
from memory_extractor import extract_memory, merge_memory_batches
from memory_learning_db import save_cost_document, update_project_budget
from memory_learning_models import CostWorkbookResult
from project_batch_ingestion import classify_document
from gemini_extractor import _structured_call, get_client

WORKFLOW_VERSION = "28.1.7.2"
LEGACY_MATERIALIZER_VERSIONS = {"28.1.1", "28.1.5", "28.1.6", "28.1.7", "28.1.7.1", "28.1.7.2"}
PROJECT_FILES_BUCKET = "nave-project-files"
MAX_SOURCE_FILES_REPAIR = 250
MAX_COST_ROWS = 2500
MAX_MEMORY_ITEMS = 90


@dataclass
class MaterializationResult:
    source_file_id: str
    project_id: str
    role: str
    status: str
    created: dict[str, int]
    warnings: list[str]

    def as_dict(self) -> dict[str, Any]:
        return {
            "source_file_id": self.source_file_id,
            "project_id": self.project_id,
            "role": self.role,
            "status": self.status,
            "created": dict(self.created),
            "warnings": list(self.warnings),
        }


def _rows(response: Any) -> list[dict[str, Any]]:
    return [dict(row) for row in (getattr(response, "data", None) or []) if isinstance(row, Mapping)]


def _normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _clip(value: Any, limit: int) -> str | None:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit] if text else None


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _safe_json(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, Mapping):
        return {str(k): _safe_json(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_safe_json(v) for v in value]
    try:
        if hasattr(value, "item"):
            return _safe_json(value.item())
    except Exception:
        pass
    return str(value)


def _source_marker(source_file: Mapping[str, Any]) -> str:
    sha = str(source_file.get("sha256") or "").strip()
    sid = str(source_file.get("id") or "").strip()
    return f"NAVE-V28.1.1:{sha or sid}"


def _download_bytes(client: Any, source_file: Mapping[str, Any]) -> bytes | None:
    bucket = str(source_file.get("storage_bucket") or "").strip()
    path = str(source_file.get("storage_path") or "").strip()
    if not bucket or not path:
        return None
    try:
        data = client.storage.from_(bucket).download(path)
        if isinstance(data, bytes):
            return data
        if isinstance(data, bytearray):
            return bytes(data)
        if hasattr(data, "read"):
            return data.read()
    except Exception:
        return None
    return None


def _table_select_one(client: Any, table: str, **filters: Any) -> dict[str, Any] | None:
    try:
        query = client.table(table).select("*")
        for key, value in filters.items():
            query = query.eq(key, value)
        result = _rows(query.limit(1).execute())
        return result[0] if result else None
    except Exception:
        return None


def _sync_project_file(client: Any, source_file_id: str, warnings: list[str]) -> bool:
    """Espelha source_files em project_files sem assumir a versão exata do schema.

    O SQL da V28.1.1 cria uma RPC idempotente que inspeciona o schema real da
    tabela project_files em produção. Se a tabela não existir ou tiver uma
    variação inesperada, a materialização especializada continua funcionando.
    """
    try:
        response = client.rpc(
            "nave_v2811_sync_project_file",
            {"p_source_file_id": source_file_id},
        ).execute()
        data = getattr(response, "data", None)
        if isinstance(data, dict):
            ok = bool(data.get("ok", True))
            if not ok:
                detail = str(data.get("detail") or data.get("reason") or "falha não especificada")
                warnings.append(f"Central de arquivos: {detail}")
            return ok
        return True
    except Exception as exc:
        warnings.append(f"Central de arquivos: {exc}")
        return False


def _parse_brl_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    try:
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            parts = text.split(",")
            if len(parts[-1]) in {1, 2}:
                text = "".join(parts[:-1]).replace(".", "") + "." + parts[-1]
            else:
                text = text.replace(",", "")
        elif text.count(".") > 1:
            parts = text.split(".")
            if len(parts[-1]) in {1, 2}:
                text = "".join(parts[:-1]) + "." + parts[-1]
            else:
                text = text.replace(".", "")
        return float(text)
    except Exception:
        return None


def _first_brl_amount(text: str) -> float | None:
    matches = re.findall(r"R\$\s*[-+]?\s*[0-9][0-9\.\s]*(?:,[0-9]{1,2})?", text or "", flags=re.I)
    for match in matches:
        parsed = _parse_brl_number(match)
        if parsed is not None and abs(parsed) < 1_000_000_000:
            return parsed
    return None


_BRIEFING_HEADINGS: list[tuple[str, str]] = [
    (r"\bobjetivo(?:s)?(?:\s+e\s+desafio)?\b", "objective"),
    (r"\bdesafio(?:s)?\b", "context"),
    (r"\bpublico\s+alvo\b|\baudiencia\b|\btarget\b", "audience"),
    (r"\bentregaveis?\b|\bentregas?\b|\bescopo\b", "deliverable"),
    (r"\bobrigatoriedades?\b|\bmandatorios?\b|\bmust\s+have\b", "mandatory"),
    (r"\brestricoes?\b|\bnao\s+pode\b", "restriction"),
    (r"\binformacoes?\s+logisticas?\b|\blogistica\b", "logistics"),
    (r"\bverba\b|\bbudget\b|\borcamento\b", "budget"),
    (r"\bkpis?\b|\bindicadores?\b|\bresultado\s+esperado\b", "kpi"),
    (r"\boperacao\b|\boperacional\b|\bfluxo\b", "operation"),
    (r"\bcomunicacao\b|\bkv\b|\bkey\s+visual\b", "communication"),
    (r"\bdesejaveis?\b|\bnice\s+to\s+have\b", "desirable"),
    (r"\bcontexto\b|\bbackground\b|\bcenario\b", "context"),
]


def _briefing_sections(text: str) -> list[dict[str, Any]]:
    source = _clean_text(text)
    if not source:
        return []

    # Mantém o mesmo comprimento aproximado do texto original para preservar
    # as posições dos recortes, mas remove acentos e caixa para localizar
    # cabeçalhos. Marcadores fracos (ex.: "KV", "operação") só viram divisão
    # quando parecem de fato um título: escritos em caixa alta ou seguidos de
    # dois-pontos. Isso evita quebrar frases como "usar KV aprovado".
    decomposed = unicodedata.normalize("NFKD", source)
    searchable = "".join(ch for ch in decomposed if not unicodedata.combining(ch)).casefold()

    strong_types = {
        "objective", "audience", "deliverable", "mandatory", "restriction",
        "logistics", "budget", "kpi", "desirable", "context",
    }
    raw_markers: list[tuple[int, int, str, str]] = []
    for pattern, req_type in _BRIEFING_HEADINGS:
        for match in re.finditer(pattern, searchable, flags=re.I):
            start, end = match.start(), match.end()
            original_match = source[start:end]
            letters = "".join(ch for ch in original_match if ch.isalpha())
            looks_upper = bool(letters) and letters.upper() == letters
            after = source[end:end + 4]
            has_colon = ":" in after
            weak_heading = has_colon or (looks_upper and len(letters) > 3)
            if req_type not in strong_types and not weak_heading:
                continue
            raw_markers.append((start, end, req_type, original_match))

    # Remove sobreposições: "OBJETIVO E DESAFIO" deve ser um único cabeçalho,
    # e não gerar uma segunda seção a partir da palavra "DESAFIO".
    raw_markers.sort(key=lambda item: (item[0], -(item[1] - item[0])))
    markers: list[tuple[int, int, str, str]] = []
    for marker in raw_markers:
        if markers and marker[0] < markers[-1][1]:
            continue
        markers.append(marker)

    sections: list[dict[str, Any]] = []
    if markers:
        for index, marker in enumerate(markers[:60]):
            start = marker[1]
            end = markers[index + 1][0] if index + 1 < len(markers) else len(source)
            piece = source[start:end].strip(" :-;,.|")
            if not piece:
                continue
            sections.append({
                "requirement_type": marker[2],
                "text": piece[:1800],
                "heading": marker[3],
            })
    else:
        sentences = re.split(r"(?<=[\.!?;])\s+|\s+[•·]\s+", source)
        for sentence in sentences:
            sentence = _clean_text(sentence)
            if len(sentence) < 18:
                continue
            norm = _normalize(sentence)
            req_type = "context"
            best = 0
            for pattern, candidate in _BRIEFING_HEADINGS:
                if re.search(pattern, norm, flags=re.I):
                    score = 3
                    if candidate in {"mandatory", "restriction", "deliverable", "objective"}:
                        score += 1
                    if score > best:
                        req_type, best = candidate, score
            sections.append({
                "requirement_type": req_type,
                "text": sentence[:1800],
                "heading": "",
            })
            if len(sections) >= 60:
                break

    if not sections:
        sections = [{"requirement_type": "context", "text": source[:1800], "heading": ""}]
    return sections


def _requirement_title(req_type: str, text: str, order: int) -> str:
    labels = {
        "objective": "Objetivo",
        "deliverable": "Entregável",
        "mandatory": "Obrigatoriedade",
        "restriction": "Restrição",
        "audience": "Público",
        "logistics": "Logística",
        "budget": "Verba",
        "kpi": "KPI / resultado esperado",
        "operation": "Operação",
        "communication": "Comunicação",
        "desirable": "Desejável",
        "context": "Contexto",
    }
    cleaned = _clean_text(text).strip(" -–—:;,.|")
    if cleaned:
        first = re.split(r"(?<=[\.!?;])\s+", cleaned)[0][:135]
        if len(first) >= 12:
            return first
    return f"{labels.get(req_type, 'Demanda')} {order}"


def _materialize_briefing(
    client: Any,
    source_file: Mapping[str, Any],
    text: str,
    warnings: list[str],
) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    existing = _table_select_one(client, "memory_briefing_documents", project_id=project_id, content_sha256=sha)
    if existing:
        return {"briefing_documents": 0, "briefing_requirements": 0}

    sections = _briefing_sections(text)
    objective = next((s["text"] for s in sections if s["requirement_type"] == "objective"), None)
    audience = next((s["text"] for s in sections if s["requirement_type"] == "audience"), None)
    payload = {
        "project_id": project_id,
        "title": str(source_file.get("file_name") or "Briefing original"),
        "file_name": str(source_file.get("file_name") or "briefing"),
        "mime_type": str(source_file.get("mime_type") or "application/octet-stream"),
        "content_sha256": sha,
        "storage_bucket": source_file.get("storage_bucket"),
        "storage_path": source_file.get("storage_path"),
        "extraction_status": "pronto",
        "requirements_count": len(sections),
        "budget_amount": _first_brl_amount(text),
        "currency": "BRL",
        "objective": _clip(objective, 4000),
        "audience": _clip(audience, 4000),
        "metadata": {
            "materialized_by": WORKFLOW_VERSION,
            "source_file_id": source_file.get("id"),
            "document_role": source_file.get("document_role"),
        },
        "diagnostic": {
            "input_mode": "flexible",
            "structured_automatically": True,
            "requirements_detected": len(sections),
        },
    }
    response = client.table("memory_briefing_documents").insert(_safe_json(payload)).execute()
    rows = _rows(response)
    if not rows:
        raise RuntimeError("o Supabase não confirmou o briefing estruturado")
    document_id = str(rows[0]["id"])

    requirement_rows: list[dict[str, Any]] = []
    for order, section in enumerate(sections, start=1):
        req_type = str(section["requirement_type"])
        content = str(section["text"])
        norm = _normalize(content)
        is_mandatory = req_type == "mandatory" or any(term in norm for term in (" deve ", " obrigatorio ", " precisa ", " necessario "))
        priority = "high" if is_mandatory or req_type in {"restriction", "objective"} else "not_informed"
        requirement_rows.append({
            "project_id": project_id,
            "briefing_document_id": document_id,
            "requirement_type": req_type,
            "title": _requirement_title(req_type, content, order),
            "description": _clip(content, 4000),
            "priority": priority,
            "mandatory": is_mandatory,
            "source_reference": str(source_file.get("file_name") or "Briefing original"),
            "source_quote": _clip(content, 1800),
            "tags": [],
            "sort_order": order,
            "adherence_status": "not_assessed",
        })
    created = 0
    for start in range(0, len(requirement_rows), 25):
        try:
            inserted = client.table("memory_briefing_requirements").insert(_safe_json(requirement_rows[start:start + 25])).execute()
            created += len(_rows(inserted))
        except Exception as exc:
            warnings.append(f"Demandas do briefing: {exc}")
            for row in requirement_rows[start:start + 25]:
                try:
                    inserted = client.table("memory_briefing_requirements").insert(_safe_json(row)).execute()
                    created += len(_rows(inserted))
                except Exception as item_exc:
                    warnings.append(f"Demanda não salva: {item_exc}")
    if created != len(sections):
        try:
            client.table("memory_briefing_documents").update({"requirements_count": created}).eq("id", document_id).execute()
        except Exception:
            pass
    return {"briefing_documents": 1, "briefing_requirements": created}


_COST_ALIASES: dict[str, tuple[str, ...]] = {
    "item_code": ("codigo", "cod", "sku", "ref", "referencia", "item code"),
    "category": ("categoria", "grupo", "secao", "centro de custo", "tipo"),
    "item_name": ("item", "descricao", "descricao item", "produto", "servico", "nome", "entrega", "escopo"),
    "description": ("detalhe", "detalhes", "observacao", "observacoes", "especificacao", "descricao completa"),
    "billing_type": ("cobranca", "faturamento", "unidade", "un", "tipo cobranca"),
    "quantity": ("qtd", "qtde", "quantidade", "qte", "qty"),
    "period": ("periodo", "dias", "diarias", "horas", "meses"),
    "unit_value": ("valor unitario", "preco unitario", "vl unit", "unitario", "unit price"),
    "base_value": ("valor base", "subtotal", "custo base", "custo"),
    "fees_value": ("fee", "honorarios", "honorario", "taxa agencia"),
    "charges_value": ("impostos", "imposto", "encargos", "taxas", "charges"),
    "client_total": ("valor total", "total cliente", "total", "preco total", "valor venda", "valor final"),
}


def _header_alias_score(cell: Any, aliases: Sequence[str]) -> int:
    norm = _normalize(cell)
    if not norm:
        return 0
    best = 0
    for alias in aliases:
        a = _normalize(alias)
        if norm == a:
            best = max(best, 4)
        elif a in norm or norm in a:
            best = max(best, 2)
    return best


def _detect_header(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    best_row: int | None = None
    best_map: dict[str, int] = {}
    best_score = 0
    for index, row in enumerate(rows[:70]):
        mapping: dict[str, int] = {}
        score = 0
        for col, cell in enumerate(row[:60]):
            matches = [(field, _header_alias_score(cell, aliases)) for field, aliases in _COST_ALIASES.items()]
            field, field_score = max(matches, key=lambda item: item[1])
            if field_score > 0 and field not in mapping:
                mapping[field] = col
                score += field_score
        if "item_name" in mapping:
            score += 4
        if any(field in mapping for field in ("client_total", "base_value", "unit_value")):
            score += 3
        if score > best_score:
            best_row, best_map, best_score = index, mapping, score
    if best_score < 6:
        return None, {}
    return best_row, best_map


def _read_spreadsheet_rows(data: bytes, file_name: str) -> list[tuple[str, list[list[Any]]]]:
    suffix = Path(file_name).suffix.lower()
    result: list[tuple[str, list[list[Any]]]] = []
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook  # type: ignore
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_vba=False)
        try:
            for sheet in workbook.worksheets[:30]:
                rows: list[list[Any]] = []
                for row in sheet.iter_rows(values_only=True):
                    values = list(row[:60])
                    rows.append(values)
                    if len(rows) >= MAX_COST_ROWS:
                        break
                result.append((str(sheet.title), rows))
        finally:
            workbook.close()
        return result
    if suffix == ".csv":
        text = data.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"  # type: ignore[attr-defined]
        reader = csv.reader(io.StringIO(text), dialect)
        rows = [list(row[:60]) for _, row in zip(range(MAX_COST_ROWS), reader)]
        return [("CSV", rows)]
    if suffix == ".xls":
        import pandas as pd  # type: ignore
        sheets = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None, dtype=object)
        for name, frame in list(sheets.items())[:30]:
            values = frame.iloc[:MAX_COST_ROWS, :60].where(frame.notna(), None).values.tolist()
            result.append((str(name), values))
        return result
    return []


def _value_at(row: Sequence[Any], mapping: Mapping[str, int], field: str) -> Any:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _looks_like_section(row: Sequence[Any]) -> str | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if not values or len(values) > 2:
        return None
    if any(_parse_brl_number(value) is not None and re.search(r"\d", value) for value in values):
        return None
    text = " - ".join(values)
    if len(text) > 120:
        return None
    uppercase_ratio = sum(ch.isupper() for ch in text if ch.isalpha()) / max(1, sum(ch.isalpha() for ch in text))
    if uppercase_ratio >= 0.7 or any(term in _normalize(text) for term in ("cenografia", "producao", "logistica", "tecnologia", "equipe", "brindes", "alimentacao", "fee")):
        return text
    return None


def _cost_items_from_data(data: bytes, file_name: str, fallback_text: str) -> tuple[list[dict[str, Any]], str | None, int | None]:
    sheets = _read_spreadsheet_rows(data, file_name)
    items: list[dict[str, Any]] = []
    first_sheet: str | None = sheets[0][0] if sheets else None
    first_header: int | None = None
    for sheet_name, rows in sheets:
        header_index, mapping = _detect_header(rows)
        if first_header is None and header_index is not None:
            first_header = header_index + 1
        current_category: str | None = None
        start = header_index + 1 if header_index is not None else 0
        for row_number, row in enumerate(rows[start:], start=start + 1):
            if len(items) >= MAX_COST_ROWS:
                break
            values = [value for value in row if value not in (None, "") and _clean_text(value)]
            if not values:
                continue
            if header_index is None:
                section = _looks_like_section(row)
                if section:
                    current_category = section[:250]
                    continue

            item_name = _clean_text(_value_at(row, mapping, "item_name")) if mapping else ""
            description = _clean_text(_value_at(row, mapping, "description")) if mapping else ""
            category = _clean_text(_value_at(row, mapping, "category")) if mapping else ""
            item_code = _clean_text(_value_at(row, mapping, "item_code")) if mapping else ""
            billing_type = _clean_text(_value_at(row, mapping, "billing_type")) if mapping else ""

            if not item_name:
                text_candidates = []
                for value in row[:25]:
                    text_value = _clean_text(value)
                    if not text_value:
                        continue
                    # Valores puramente numéricos não são bons nomes de item.
                    if _parse_brl_number(value) is not None and re.fullmatch(r"[\d\s\.,R$%\-]+", text_value):
                        continue
                    if len(text_value) >= 3:
                        text_candidates.append(text_value)
                if text_candidates:
                    item_name = max(text_candidates, key=lambda item: min(len(item), 120))[:500]

            numeric_values = [_parse_brl_number(value) for value in row]
            numeric_values = [value for value in numeric_values if value is not None]
            quantity = _parse_brl_number(_value_at(row, mapping, "quantity")) if mapping else None
            period = _parse_brl_number(_value_at(row, mapping, "period")) if mapping else None
            unit_value = _parse_brl_number(_value_at(row, mapping, "unit_value")) if mapping else None
            base_value = _parse_brl_number(_value_at(row, mapping, "base_value")) if mapping else None
            fees_value = _parse_brl_number(_value_at(row, mapping, "fees_value")) if mapping else None
            charges_value = _parse_brl_number(_value_at(row, mapping, "charges_value")) if mapping else None
            client_total = _parse_brl_number(_value_at(row, mapping, "client_total")) if mapping else None
            if header_index is None and client_total is None and numeric_values:
                # Sem cabeçalho confiável, o último número monetário é guardado
                # como total apenas se a linha também contiver descrição textual.
                client_total = numeric_values[-1] if item_name else None

            if not item_name:
                continue
            # Evita transformar cabeçalhos repetidos em itens.
            if _normalize(item_name) in {_normalize(alias) for aliases in _COST_ALIASES.values() for alias in aliases}:
                continue

            has_value = any(value is not None for value in (unit_value, base_value, fees_value, charges_value, client_total))
            item_status = "included" if has_value else "no_value"
            estimate_type = "quoted" if has_value else "no_value"
            items.append({
                "source_sheet": sheet_name,
                "source_row": row_number,
                "item_code": item_code or None,
                "category": (category or current_category or None),
                "item_name": item_name[:500],
                "description": _clip(description, 4000),
                "billing_type": billing_type or None,
                "quantity": quantity,
                "period": period,
                "unit_value": unit_value,
                "base_value": base_value,
                "fees_value": fees_value,
                "charges_value": charges_value,
                "client_total": client_total,
                "item_status": item_status,
                "estimate_type": estimate_type,
                "flags": ([] if header_index is not None else ["estrutura_inferida"]),
                "raw_data": {
                    "materialized_by": WORKFLOW_VERSION,
                    "source_cells": [_safe_json(value) for value in row[:30]],
                    "header_detected": header_index is not None,
                },
            })
    if not items and _clean_text(fallback_text):
        items.append({
            "source_sheet": first_sheet or "Documento",
            "source_row": 1,
            "item_code": None,
            "category": "Não classificado",
            "item_name": "Conteúdo da planilha de custos",
            "description": _clip(fallback_text, 4000),
            "billing_type": None,
            "quantity": None,
            "period": None,
            "unit_value": None,
            "base_value": None,
            "fees_value": None,
            "charges_value": None,
            "client_total": None,
            "item_status": "no_value",
            "estimate_type": "no_value",
            "flags": ["estrutura_inferida", "revisar_se_necessario"],
            "raw_data": {"materialized_by": WORKFLOW_VERSION, "fallback": True},
        })
    return items, first_sheet, first_header


def _materialize_costs(
    client: Any,
    source_file: Mapping[str, Any],
    text: str,
    data: bytes | None,
    warnings: list[str],
) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    existing = _table_select_one(client, "memory_cost_documents", project_id=project_id, content_sha256=sha)
    if existing:
        return {"cost_documents": 0, "cost_items": 0}
    if data is None:
        data = _download_bytes(client, source_file)
    if data is None:
        raise RuntimeError("a planilha original não pôde ser recuperada do Storage")

    items, first_sheet, header_row = _cost_items_from_data(data, str(source_file.get("file_name") or "planilha.xlsx"), text)
    totals = [float(item["client_total"]) for item in items if item.get("client_total") is not None]
    payload = {
        "project_id": project_id,
        "title": str(source_file.get("file_name") or "Planilha de custos"),
        "file_name": str(source_file.get("file_name") or "planilha"),
        "mime_type": str(source_file.get("mime_type") or "application/octet-stream"),
        "sheet_name": first_sheet,
        "header_row": header_row,
        "content_sha256": sha,
        "storage_bucket": source_file.get("storage_bucket"),
        "storage_path": source_file.get("storage_path"),
        "extraction_status": "pronto",
        "total_items": len(items),
        "included_items": sum(1 for item in items if item["item_status"] == "included"),
        "optional_items": 0,
        "pending_items": sum(1 for item in items if item["item_status"] == "no_value"),
        "total_base": sum(float(item["base_value"]) for item in items if item.get("base_value") is not None) or None,
        "fees_total": sum(float(item["fees_value"]) for item in items if item.get("fees_value") is not None) or None,
        "charges_total": sum(float(item["charges_value"]) for item in items if item.get("charges_value") is not None) or None,
        "client_total": sum(totals) if totals else None,
        "currency": "BRL",
        "macros_present": Path(str(source_file.get("file_name") or "")).suffix.lower() == ".xlsm",
        "metadata": {
            "materialized_by": WORKFLOW_VERSION,
            "source_file_id": source_file.get("id"),
            "flexible_input": True,
        },
        "diagnostic": {
            "items_detected": len(items),
            "header_row_inferred": header_row,
            "requires_exact_template": False,
        },
        "raw_data": {"document_role": source_file.get("document_role")},
    }
    inserted = _rows(client.table("memory_cost_documents").insert(_safe_json(payload)).execute())
    if not inserted:
        raise RuntimeError("o Supabase não confirmou a planilha de custos estruturada")
    doc_id = str(inserted[0]["id"])

    item_payloads = []
    for item in items:
        row = dict(item)
        row["project_id"] = project_id
        row["cost_document_id"] = doc_id
        item_payloads.append(row)
    created = 0
    for start in range(0, len(item_payloads), 35):
        chunk = item_payloads[start:start + 35]
        try:
            created += len(_rows(client.table("memory_cost_items").insert(_safe_json(chunk)).execute()))
        except Exception as exc:
            warnings.append(f"Itens de custo: {exc}")
            for row in chunk:
                try:
                    created += len(_rows(client.table("memory_cost_items").insert(_safe_json(row)).execute()))
                except Exception as item_exc:
                    warnings.append(f"Linha de custo não salva: {item_exc}")
    if created != len(items):
        try:
            client.table("memory_cost_documents").update({
                "total_items": created,
                "included_items": min(created, payload["included_items"]),
            }).eq("id", doc_id).execute()
        except Exception:
            pass
    return {"cost_documents": 1, "cost_items": created}


_MEMORY_SECTION_TERMS: dict[str, tuple[str, ...]] = {
    "strategy": ("estrategia", "conceito", "insight", "objetivo", "posicionamento", "big idea", "territorio"),
    "scenography": ("cenografia", "ambientacao", "ambiente", "palco", "estande", "stand", "layout", "arquitetura", "mobiliario"),
    "activations": ("ativacao", "experiencia", "dinamica", "game", "jogo", "photo op", "photoop", "interacao", "mecanica"),
    "gifts": ("brinde", "press kit", "presskit", "kit", "gift", "presente", "merch", "produto promocional"),
    "journey_operation": ("jornada", "operacao", "fluxo", "fila", "credenciamento", "promotor", "staff", "equipe", "logistica"),
    "communication": ("comunicacao", "kv", "key visual", "identidade visual", "convite", "sinalizacao", "social", "peca"),
    "content_agenda": ("agenda", "programacao", "conteudo", "palestra", "painel", "show", "workshop", "cronograma"),
    "partners_sponsorship": ("parceiro", "fornecedor", "patrocinio", "patrocinador", "cota", "sponsor", "apoio"),
    "pr_esg_legacy": ("imprensa", "pr ", "relacoes publicas", "sustentabilidade", "esg", "legado", "impacto social"),
}


def _split_memory_chunks(text: str) -> list[str]:
    source = _clean_text(text)
    if not source:
        return []
    sentences = [
        _clean_text(item)
        for item in re.split(r"(?<=[\.!?;])\s+|\s+[•·]\s+", source)
        if len(_clean_text(item)) >= 18
    ]
    chunks: list[str] = []
    buffer: list[str] = []
    length = 0
    for sentence in sentences:
        buffer.append(sentence)
        length += len(sentence)
        if length >= 380 or len(buffer) >= 3:
            chunks.append(" ".join(buffer)[:1500])
            buffer, length = [], 0
        if len(chunks) >= MAX_MEMORY_ITEMS:
            break
    if buffer and len(chunks) < MAX_MEMORY_ITEMS:
        chunks.append(" ".join(buffer)[:1500])
    if not chunks:
        chunks = [source[:1500]]
    return chunks


def _memory_section_for_text(text: str) -> tuple[str, float]:
    norm = f" {_normalize(text)} "
    scores: dict[str, int] = {}
    for section, terms in _MEMORY_SECTION_TERMS.items():
        scores[section] = sum(1 for term in terms if f" {_normalize(term)} " in norm or _normalize(term) in norm)
    section, score = max(scores.items(), key=lambda item: item[1])
    if score <= 0:
        return "strategy", 0.52
    return section, min(0.95, 0.58 + score * 0.08)


def _memory_title(section: str, text: str, order: int) -> str:
    labels = {
        "strategy": "Estratégia / conceito",
        "scenography": "Cenografia / ambiente",
        "activations": "Ativação / experiência",
        "gifts": "Brinde / material",
        "journey_operation": "Jornada / operação",
        "communication": "Comunicação",
        "content_agenda": "Conteúdo / agenda",
        "partners_sponsorship": "Parceiros / fornecedores",
        "pr_esg_legacy": "PR / ESG / legado",
    }
    first = re.split(r"(?<=[\.!?;])\s+", _clean_text(text))[0][:140]
    if len(first) >= 16:
        return first
    return f"{labels.get(section, 'Conteúdo')} {order}"


def _materialize_presentation(
    client: Any,
    source_file: Mapping[str, Any],
    text: str,
    warnings: list[str],
) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    existing = _table_select_one(client, "memory_documents", project_id=project_id, content_sha256=sha)
    if existing:
        return {"memory_documents": 0, "memory_items": 0}
    role = str(source_file.get("document_role") or "proposal_presentation")
    chunks = _split_memory_chunks(text)
    strategic_summary = next((chunk for chunk in chunks if _memory_section_for_text(chunk)[0] == "strategy"), chunks[0] if chunks else None)
    creative = next((chunk for chunk in chunks if "conceito" in _normalize(chunk) or "big idea" in _normalize(chunk)), None)
    payload = {
        "project_id": project_id,
        "title": str(source_file.get("file_name") or "Apresentação"),
        "file_name": str(source_file.get("file_name") or "apresentacao"),
        "mime_type": str(source_file.get("mime_type") or "application/octet-stream"),
        "version_label": "Importação de projeto",
        "document_status": "sent_to_client",
        "page_count": source_file.get("page_count"),
        "rendered_pages_count": 0,
        "items_count": len(chunks),
        "visual_crops_count": 0,
        "content_sha256": sha,
        "storage_bucket": source_file.get("storage_bucket"),
        "storage_path": source_file.get("storage_path"),
        "extraction_status": "pronto",
        "strategic_summary": _clip(strategic_summary, 6000),
        "creative_concept": _clip(creative, 6000),
        "raw_data": {
            "materialized_by": WORKFLOW_VERSION,
            "source_file_id": source_file.get("id"),
            "document_role": role,
            "visuals_not_inferred": True,
        },
    }
    inserted = _rows(client.table("memory_documents").insert(_safe_json(payload)).execute())
    if not inserted:
        raise RuntimeError("o Supabase não confirmou a apresentação estruturada")
    document_id = str(inserted[0]["id"])

    item_payloads: list[dict[str, Any]] = []
    for order, chunk in enumerate(chunks, start=1):
        section, confidence = _memory_section_for_text(chunk)
        item_payloads.append({
            "project_id": project_id,
            "document_id": document_id,
            "page_id": None,
            "source_page": 1,
            "section_key": section,
            "item_type": "Conteúdo",
            "title": _memory_title(section, chunk, order),
            "summary": _clip(chunk, 1800),
            "description": _clip(chunk, 6000),
            "item_status": "Não identificado",
            "tags": [],
            "objectives": [],
            "audiences": [],
            "mechanics": [],
            "technologies": [],
            "confidence": confidence,
            "evidence": _clip(chunk, 1800),
            "sort_order": order,
            "raw_data": {
                "materialized_by": WORKFLOW_VERSION,
                "source_file_id": source_file.get("id"),
                "source_file": source_file.get("file_name"),
                "source_page_inferred": True,
            },
        })
    created = 0
    for start in range(0, len(item_payloads), 30):
        chunk = item_payloads[start:start + 30]
        try:
            created += len(_rows(client.table("memory_items").insert(_safe_json(chunk)).execute()))
        except Exception as exc:
            warnings.append(f"Conteúdos da apresentação: {exc}")
            for row in chunk:
                try:
                    created += len(_rows(client.table("memory_items").insert(_safe_json(row)).execute()))
                except Exception as item_exc:
                    warnings.append(f"Conteúdo não salvo: {item_exc}")
    if created != len(item_payloads):
        try:
            client.table("memory_documents").update({"items_count": created}).eq("id", document_id).execute()
        except Exception:
            pass
    return {"memory_documents": 1, "memory_items": created}


def _materialize_feedback(
    client: Any,
    source_file: Mapping[str, Any],
    text: str,
) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    mime_type = str(source_file.get("mime_type") or "").casefold()
    file_name = str(source_file.get("file_name") or "arquivo")
    if _clip(text, 12000):
        content = _clip(text, 12000)
    elif mime_type.startswith("image/"):
        content = f"Feedback visual anexado: {file_name}. Conteúdo textual não extraído automaticamente nesta etapa."
    else:
        content = f"Arquivo de feedback: {file_name}"
    marker = _source_marker(source_file)
    try:
        response = (
            client.table("memory_feedback_entries")
            .select("id,internal_interpretation")
            .eq("project_id", project_id)
            .execute()
        )
        for row in _rows(response):
            if marker in str(row.get("internal_interpretation") or ""):
                return {"feedback_entries": 0}
    except Exception:
        pass
    norm = _normalize(content)
    sentiment = "mixed" if any(term in norm for term in ("aprov", "gost", "positivo")) and any(term in norm for term in ("ajust", "alter", "nao ", "negativ")) else "neutral"
    if any(term in norm for term in ("ajust", "alter", "revis")):
        stage = "revision"
    elif any(term in norm for term in ("aprov", "cliente", "apresent")):
        stage = "presentation"
    else:
        stage = "not_informed"
    theme = "budget" if any(term in norm for term in ("orcamento", "verba", "custo", "valor")) else "other"
    payload = {
        "project_id": project_id,
        "source_type": "not_informed",
        "process_stage": stage,
        "theme": theme,
        "sentiment": sentiment,
        "original_feedback": content,
        "internal_interpretation": (
            f"{marker} · Arquivo visual preservado como evidência; sem inferir texto, autor ou decisão."
            if mime_type.startswith("image/") and not _clip(text, 12000)
            else f"{marker} · Materializado automaticamente a partir do arquivo, sem inventar autor ou decisão."
        ),
        "action_taken": None,
        "confidence_level": "incomplete",
    }
    inserted = _rows(client.table("memory_feedback_entries").insert(_safe_json(payload)).execute())
    return {"feedback_entries": len(inserted)}


def _append_source(existing: str | None, marker: str, file_name: str, text: str) -> str:
    current = str(existing or "").strip()
    if marker in current:
        return current
    addition = f"[{marker}] {file_name}: {_clip(text, 3500) or 'documento anexado'}"
    return (current + "\n\n" + addition).strip()[:12000]


def _materialize_report(client: Any, source_file: Mapping[str, Any], text: str) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    marker = _source_marker(source_file)
    existing = _table_select_one(client, "memory_project_outcomes", project_id=project_id)
    file_name = str(source_file.get("file_name") or "Relatório")
    if existing:
        payload = {
            "result_context": _append_source(existing.get("result_context"), marker, file_name, text),
            "execution_notes": _append_source(existing.get("execution_notes"), marker, file_name, text),
        }
        # Preserva qualquer status/resultado já confirmado. Só melhora a origem
        # quando ainda não havia informação explícita.
        if str(existing.get("information_source") or "not_informed") == "not_informed":
            payload["information_source"] = "document"
        client.table("memory_project_outcomes").update(_safe_json(payload)).eq("project_id", project_id).execute()
        return {"project_outcomes": 0}
    payload = {
        "project_id": project_id,
        "process_type": "not_informed",
        "commercial_result": "in_evaluation",
        "proposal_result": "not_informed",
        "execution_result": "not_informed",
        "result_reasons": [],
        "result_context": _append_source(None, marker, file_name, text),
        "execution_notes": _append_source(None, marker, file_name, text),
        "currency": "BRL",
        "confidence_level": "incomplete",
        "information_source": "document",
    }
    inserted = _rows(client.table("memory_project_outcomes").insert(_safe_json(payload)).execute())
    return {"project_outcomes": len(inserted)}


def _mark_source_file(
    client: Any,
    source_file_id: str,
    *,
    status: str,
    notes: str | None,
    created: Mapping[str, int] | None = None,
) -> None:
    payload = {
        "processing_status": status,
        "processing_notes": _clip(notes, 4000),
    }
    try:
        existing = _table_select_one(client, "source_files", id=source_file_id) or {}
        metadata = dict(existing.get("metadata") or {}) if isinstance(existing.get("metadata"), Mapping) else {}
        metadata.update({
            "materialized_by": WORKFLOW_VERSION,
            "workspace_created": dict(created or {}),
        })
        payload["metadata"] = metadata
    except Exception:
        pass
    try:
        client.table("source_files").update(_safe_json(payload)).eq("id", source_file_id).execute()
    except Exception:
        pass



PROJECT_FILE_ROLE_BY_DOCUMENT_ROLE = {
    "briefing_original": "briefing_original",
    "detailed_costs": "cost_sheet",
    "preliminary_budget": "cost_sheet",
    "proposal_presentation": "project_document",
    "final_presentation": "final_presentation",
    "feedback_approval": "feedback",
    "post_event_report": "post_execution_report",
    "supplier_reference": "supplier_reference",
    "complementary_document": "project_document",
}


def _ai_settings() -> tuple[str | None, str]:
    api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    model = os.getenv("GEMINI_MODEL") or "gemini-3.5-flash-lite"
    try:
        import streamlit as st
        api_key = str(st.secrets.get("GEMINI_API_KEY") or st.secrets.get("GOOGLE_API_KEY") or api_key or "").strip() or None
        model = str(st.secrets.get("GEMINI_MODEL") or model).strip() or model
    except Exception:
        pass
    return api_key, model


def _legacy_marker(row: Mapping[str, Any], field: str) -> str:
    payload = row.get(field)
    if not isinstance(payload, Mapping):
        return ""
    return str(payload.get("materialized_by") or "").strip()


def _delete_legacy_specialized_rows(client: Any, source_file: Mapping[str, Any]) -> dict[str, int]:
    """Remove a interpretação estruturada do mesmo arquivo antes do reprocessamento.

    A ação só é chamada pelo fluxo explícito de reprocessamento. O arquivo original em
    ``source_files``/Storage não é removido. A identidade por projeto + SHA-256 garante
    que apenas a materialização daquele mesmo documento seja substituída.
    """
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    deleted = {"briefing": 0, "cost": 0, "presentation": 0}
    specs = [
        ("memory_briefing_documents", "briefing"),
        ("memory_cost_documents", "cost"),
        ("memory_documents", "presentation"),
    ]
    if not project_id or not sha:
        return deleted

    for table, key in specs:
        try:
            rows = _rows(
                client.table(table).select("*")
                .eq("project_id", project_id)
                .eq("content_sha256", sha)
                .execute()
            )
        except Exception:
            rows = []
        for row in rows:
            # Reprocessar significa substituir a leitura estruturada, não o original.
            # As tabelas filhas possuem cascade para o documento estruturado.
            try:
                client.table(table).delete().eq("id", row.get("id")).execute()
                deleted[key] += 1
            except Exception:
                pass
    return deleted


def _ensure_project_file_role(client: Any, source_file: Mapping[str, Any], warnings: list[str]) -> None:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    role = str(source_file.get("document_role") or "complementary_document")
    desired = PROJECT_FILE_ROLE_BY_DOCUMENT_ROLE.get(role, "project_document")
    if not project_id or not sha:
        return
    try:
        existing = _rows(
            client.table("project_files").select("*")
            .eq("project_id", project_id)
            .eq("content_sha256", sha)
            .execute()
        )
    except Exception as exc:
        warnings.append(f"Central de arquivos: {exc}")
        return

    for row in existing:
        if str(row.get("file_role") or "") == desired:
            # Mesmo quando o papel físico já está correto, sincronizamos a função
            # semântica original. Isso permite distinguir proposta de documento
            # genérico sem alterar o CHECK legado de project_files.
            metadata = dict(row.get("metadata") or {})
            wanted_metadata = {
                **metadata,
                "source_file_id": source_file.get("id"),
                "semantic_role_sync": WORKFLOW_VERSION,
                "document_role": role,
                "source_processing_status": source_file.get("processing_status"),
            }
            if wanted_metadata != metadata:
                try:
                    client.table("project_files").update({
                        "title": str(source_file.get("file_name") or row.get("title") or "Arquivo"),
                        "metadata": _safe_json(wanted_metadata),
                    }).eq("id", row.get("id")).execute()
                except Exception as exc:
                    warnings.append(f"Central de arquivos: não foi possível sincronizar metadados ({exc})")
            return
    if existing:
        row = existing[0]
        try:
            client.table("project_files").update({
                "file_role": desired,
                "title": str(source_file.get("file_name") or row.get("title") or "Arquivo"),
                "metadata": _safe_json({
                    **(row.get("metadata") or {}),
                    "source_file_id": source_file.get("id"),
                    "semantic_role_sync": WORKFLOW_VERSION,
                    "document_role": role,
                    "source_processing_status": source_file.get("processing_status"),
                }),
            }).eq("id", row.get("id")).execute()
            return
        except Exception as exc:
            warnings.append(f"Central de arquivos: não foi possível ajustar o papel ({exc})")
            return

    try:
        versions = _rows(
            client.table("project_files").select("version_number")
            .eq("project_id", project_id).eq("file_role", desired).execute()
        )
        version = max([int(r.get("version_number") or 0) for r in versions] or [0]) + 1
        client.table("project_files").insert({
            "project_id": project_id,
            "file_role": desired,
            "title": str(source_file.get("file_name") or "Arquivo"),
            "file_name": str(source_file.get("file_name") or "arquivo"),
            "mime_type": source_file.get("mime_type"),
            "file_size_bytes": source_file.get("file_size_bytes"),
            "content_sha256": sha,
            "version_number": version,
            "storage_bucket": str(source_file.get("storage_bucket") or PROJECT_FILES_BUCKET),
            "storage_path": str(source_file.get("storage_path") or ""),
            "metadata": _safe_json({
                "source_file_id": source_file.get("id"),
                "semantic_role_sync": WORKFLOW_VERSION,
                "document_role": role,
                "source_processing_status": source_file.get("processing_status"),
            }),
            "is_current": True,
            "is_archived": False,
        }).execute()
    except Exception as exc:
        warnings.append(f"Central de arquivos: {exc}")



def _repair_financial_source_role(
    client: Any,
    source_file: Mapping[str, Any],
    warnings: list[str],
) -> dict[str, Any]:
    """Corrige planilhas financeiras que versões anteriores rotularam pelo texto interno.

    O caso real do Planeja 27 foi salvo como ``post_event_report`` porque uma linha
    de escopo mencionava pós-evento/relatório final. Antes de decidir quais arquivos
    reprocessar, voltamos a classificar planilhas com a regra documental da V28.1.7.
    """
    row = dict(source_file)
    file_name = str(row.get("file_name") or "")
    if Path(file_name).suffix.casefold() not in {".xlsx", ".xlsm", ".xls", ".csv"}:
        return row
    predicted, confidence, reasons = classify_document(
        file_name,
        str(row.get("text_excerpt") or ""),
    )

    # A classificação histórica pode ter sido contaminada pelo próprio escopo
    # da planilha (ex.: uma linha chamada pós-evento). Quando isso acontecer,
    # validamos a identidade financeira pela estrutura real do workbook antes
    # de excluir o arquivo da rodada de reprocessamento.
    if predicted not in {"detailed_costs", "preliminary_budget"}:
        workbook_bytes = _download_bytes(client, row)
        if workbook_bytes:
            try:
                parsed = parse_cost_workbook(file_name, workbook_bytes)
                structural_kind = str((parsed.metadata or {}).get("cost_kind") or "")
                if structural_kind in {"detailed_costs", "preliminary_budget"}:
                    predicted = structural_kind
                    confidence = 0.99
                    reasons = ["estrutura financeira confirmada pelo workbook"]
            except Exception:
                pass
    if predicted not in {"detailed_costs", "preliminary_budget"}:
        return row
    current = str(row.get("document_role") or "complementary_document")
    row["document_role"] = predicted
    row["role_confidence"] = confidence
    row["role_reasons"] = reasons
    if current == predicted:
        return row
    source_id = str(row.get("id") or "")
    if not source_id:
        return row
    try:
        client.table("source_files").update({
            "document_role": predicted,
            "role_confidence": confidence,
            "role_reasons": reasons,
        }).eq("id", source_id).execute()
        warnings.append(
            f"Papel documental corrigido automaticamente: {file_name} · {current} → {predicted}."
        )
        _ensure_project_file_role(client, row, warnings)
    except Exception as exc:
        warnings.append(f"Não foi possível sincronizar o papel financeiro de {file_name}: {exc}")
    return row




def _set_source_document_role(
    client: Any,
    source_file: Mapping[str, Any],
    *,
    role: str,
    confidence: float,
    reasons: Sequence[str],
    warnings: list[str],
) -> dict[str, Any]:
    row = dict(source_file)
    current = str(row.get("document_role") or "complementary_document")
    if not row.get("id") or current == role:
        row["document_role"] = role
        row["role_confidence"] = confidence
        row["role_reasons"] = list(reasons)
        return row
    try:
        client.table("source_files").update({
            "document_role": role,
            "role_confidence": float(confidence),
            "role_reasons": list(reasons)[:12],
        }).eq("id", row.get("id")).execute()
        row["document_role"] = role
        row["role_confidence"] = confidence
        row["role_reasons"] = list(reasons)
        warnings.append(
            f"Papel documental corrigido automaticamente: {row.get('file_name') or 'arquivo'} · {current} → {role}."
        )
        _ensure_project_file_role(client, row, warnings)
    except Exception as exc:
        warnings.append(
            f"Não foi possível sincronizar o papel de {row.get('file_name') or 'arquivo'}: {exc}"
        )
    return row


def _repair_general_source_role(
    client: Any,
    source_file: Mapping[str, Any],
    warnings: list[str],
) -> dict[str, Any]:
    """Reclassifica papéis semanticamente fortes sem transformar todo PDF em proposta.

    A correção é conservadora: troca automática só ocorre quando o papel anterior
    era genérico, ou quando um não-imagem foi marcado como feedback apesar de a
    leitura documental apontar com alta confiança para briefing/apresentação.
    """
    row = dict(source_file)
    file_name = str(row.get("file_name") or "")
    extension = Path(file_name).suffix.casefold()
    if extension in {".xlsx", ".xlsm", ".xls", ".csv"}:
        return _repair_financial_source_role(client, row, warnings)

    predicted, confidence, reasons = classify_document(
        file_name,
        str(row.get("text_excerpt") or ""),
    )
    current = str(row.get("document_role") or "complementary_document")
    if predicted == "complementary_document":
        return row
    if extension in {".jpg", ".jpeg", ".png", ".webp"}:
        if current == "complementary_document" and predicted == "feedback_approval" and confidence >= 0.62:
            return _set_source_document_role(
                client, row, role=predicted, confidence=confidence,
                reasons=reasons, warnings=warnings,
            )
        return row

    safe_to_replace = (
        current == "complementary_document" and confidence >= 0.66
    ) or (
        current == "feedback_approval"
        and predicted in {"briefing_original", "proposal_presentation", "final_presentation"}
        and confidence >= 0.78
    )
    if not safe_to_replace or predicted == current:
        return row
    return _set_source_document_role(
        client, row, role=predicted, confidence=confidence,
        reasons=reasons, warnings=warnings,
    )


def _repair_bundle_topology(
    client: Any,
    source_files: Sequence[Mapping[str, Any]],
    warnings: list[str],
) -> list[dict[str, Any]]:
    """Usa a composição do lote apenas como fallback para papéis ausentes.

    Caso típico: um lote contém uma planilha de custos, um print de feedback, um
    Word de briefing e um PDF de proposta com nome genérico. Se briefing ou
    apresentação ficaram sem papel específico, a NAVE corrige apenas quando há
    um único candidato plausível, preservando a revisão humana como autoridade.
    """
    rows = [dict(row) for row in source_files]
    roles = {str(row.get("document_role") or "") for row in rows}

    if "briefing_original" not in roles:
        candidates = [
            row for row in rows
            if Path(str(row.get("file_name") or "")).suffix.casefold() in {".docx", ".txt", ".md"}
            and str(row.get("document_role") or "") == "complementary_document"
            and len(str(row.get("text_excerpt") or "").strip()) >= 350
        ]
        if len(candidates) == 1:
            fixed = _set_source_document_role(
                client, candidates[0], role="briefing_original", confidence=0.72,
                reasons=["único documento textual substancial do lote e briefing ausente"],
                warnings=warnings,
            )
            rows = [fixed if str(row.get("id")) == str(fixed.get("id")) else row for row in rows]
            roles.add("briefing_original")

    if not ({"proposal_presentation", "final_presentation"} & roles):
        bundle_has_briefing_and_cost = (
            "briefing_original" in roles
            and bool({"detailed_costs", "preliminary_budget"} & roles)
        )
        candidates = [
            row for row in rows
            if bundle_has_briefing_and_cost
            and Path(str(row.get("file_name") or "")).suffix.casefold() == ".pdf"
            and str(row.get("document_role") or "") in {"complementary_document", "feedback_approval"}
            and int(row.get("page_count") or 0) >= 4
        ]
        if len(candidates) == 1:
            fixed = _set_source_document_role(
                client, candidates[0], role="proposal_presentation", confidence=0.70,
                reasons=["único PDF multipágina do lote e apresentação ausente"],
                warnings=warnings,
            )
            rows = [fixed if str(row.get("id")) == str(fixed.get("id")) else row for row in rows]

    return rows


def _source_document(source_file: Mapping[str, Any], data: bytes):
    docs = prepare_input_documents([(
        str(source_file.get("file_name") or "arquivo"),
        data,
        str(source_file.get("mime_type") or "application/octet-stream"),
    )])
    if not docs:
        raise RuntimeError("o arquivo original não pôde ser preparado para leitura")
    return docs[0]


def _materialize_briefing_semantic(client: Any, source_file: Mapping[str, Any], data: bytes, warnings: list[str]) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    existing = _table_select_one(client, "memory_briefing_documents", project_id=project_id, content_sha256=sha)
    if existing:
        return {"briefing_documents": 0, "briefing_requirements": 0}
    api_key, model = _ai_settings()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY não está configurada para a leitura especializada do briefing")
    doc = _source_document(source_file, data)
    extraction = analyze_briefing_document(doc, api_key=api_key, model=model)
    payload = {
        "project_id": project_id,
        "title": extraction.title or str(source_file.get("file_name") or "Briefing original"),
        "file_name": str(source_file.get("file_name") or "briefing"),
        "mime_type": str(source_file.get("mime_type") or doc.mime_type),
        "content_sha256": sha,
        "storage_bucket": source_file.get("storage_bucket"),
        "storage_path": source_file.get("storage_path"),
        "extraction_status": "pronto",
        "requirements_count": len(extraction.requirements),
        "budget_amount": extraction.budget_amount,
        "currency": extraction.currency or "BRL",
        "objective": _clip(extraction.objective, 4000),
        "audience": _clip(extraction.audience, 4000),
        "metadata": {
            "materialized_by": WORKFLOW_VERSION,
            "source_file_id": source_file.get("id"),
            "document_role": source_file.get("document_role"),
            "event_date": extraction.event_date,
            "venue": extraction.venue,
            "audience_quantity": extraction.audience_quantity,
            "client_brand": extraction.client_brand,
            "event_name": extraction.event_name,
        },
        "diagnostic": {"warnings": extraction.warnings, "semantic_pipeline": True},
    }
    inserted = _rows(client.table("memory_briefing_documents").insert(_safe_json(payload)).execute())
    if not inserted:
        raise RuntimeError("o Supabase não confirmou o briefing estruturado")
    document_id = str(inserted[0]["id"])
    rows = []
    for order, req in enumerate(extraction.requirements, start=1):
        rows.append({
            "project_id": project_id,
            "briefing_document_id": document_id,
            "requirement_type": req.requirement_type,
            "title": _clip(req.title, 500) or f"Demanda {order}",
            "description": _clip(req.description, 5000),
            "priority": req.priority,
            "mandatory": bool(req.mandatory),
            "source_reference": _clip(req.source_reference, 500),
            "source_quote": _clip(req.source_quote, 1800),
            "tags": list(req.tags or []),
            "sort_order": order,
            "adherence_status": "not_assessed",
        })
    created = 0
    for start in range(0, len(rows), 30):
        created += len(_rows(client.table("memory_briefing_requirements").insert(_safe_json(rows[start:start+30])).execute()))
    if extraction.budget_amount is not None:
        try:
            update_project_budget(client, project_id=project_id, budget_amount=extraction.budget_amount, currency=extraction.currency or "BRL")
        except Exception as exc:
            warnings.append(f"Budget do briefing: {exc}")
    return {"briefing_documents": 1, "briefing_requirements": created}


def _scoped_page_text(page_text: str, pattern: str, limit: int = 520) -> str | None:
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(page_text or "").splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return None
    try:
        matcher = re.compile(pattern, flags=re.IGNORECASE)
    except re.error:
        return None
    for index, line in enumerate(lines):
        if not matcher.search(line):
            continue
        selected = [line]
        for next_line in lines[index + 1:index + 3]:
            if len(" ".join(selected + [next_line])) > limit:
                break
            selected.append(next_line)
        return _clip(" ".join(selected), limit)
    return None


def _semantic_pattern(text: str) -> str | None:
    if "photo op" in text or "ponto de foto" in text:
        return r"photo[\s-]?op|ponto de foto"
    if "quick massage" in text or "ilha de massagem" in text or "massagem" in text:
        return r"quick massage|ilha de massagem|massagem"
    if "bar de cafe" in text:
        return r"bar de caf[eé]s?|caf[eé]"
    if "bacio di latte" in text:
        return r"bacio di latte"
    if "palco" in text:
        return r"\bpalco\b"
    if "totem" in text:
        return r"\btotem\b"
    if "kit de boas vindas" in text or "kit boas vindas" in text:
        return r"kit de boas[- ]vindas|kit boas[- ]vindas"
    if "camiseta" in text:
        return r"camiseta"
    if "mala" in text or "mochila" in text:
        return r"mala|mochila"
    if "jantar tematico" in text:
        return r"jantar tem[aá]tico"
    if "banda" in text or "musica ao vivo" in text:
        return r"banda|m[uú]sica ao vivo"
    return None


def _sanitize_memory_items(
    items: list[dict[str, Any]],
    page_inventory: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    pages = {
        int(row.get("page_number") or 0): row
        for row in (page_inventory or [])
        if int(row.get("page_number") or 0) > 0
    }
    food_terms = (
        "risotto", "risoto", "panna cotta", "cannoli", "ossobuco", "polenta",
        "insalata", "antipasti", "zabaione", "minestrone", "cardapio", "menu",
    )
    schedule_terms = (
        "check in", "almoco", "jantar", "coffee", "plenaria", "check out", "coquetel",
        "happy hour", "dia 0", "dia 1", "dia 2", "dia 3", "dia 4",
    )
    gift_terms = (
        "brinde", "gift", "press kit", "presskit", "camiseta", "mochila", "mala para esportes",
        "credencial", "sacola", "kit boas vindas", "kit de boas vindas",
    )
    activation_terms = (
        "photo op", "ponto de foto", "quick massage", "ilha de massagem", "bar de cafe",
        "bacio di latte", "ativacao", "experiencia",
    )
    scenography_terms = (
        "palco", "totem led", "totem", "cenografia", "implantacao", "planta", "led 14x4",
        "led 6x2", "housemix", "backdrop",
    )
    venue_terms = (
        "bourbon atibaia", "hotel bourbon", "ballroom garden", "sala londrina", "board room",
        "orquidea", "tulipa", "montagem dias",
    )
    artistic_terms = ("banda", "studio 4", "musica ao vivo", "palestrante", "mestre de cerimonias")
    strategy_terms = (
        "brief recap", "objetivo", "estrategia", "conceito", "racional", "insight", "narrativa",
        "direcional", "territorio criativo", "premissa",
    )
    generic_strategy = re.compile(r"^estrategia(?: e conceito)?(?: slide)? \d+$")

    cleaned: list[dict[str, Any]] = []
    for item in items:
        row = dict(item)
        text = _normalize(" ".join(str(row.get(k) or "") for k in ("title", "item_type", "summary", "description", "evidence")))
        title_norm = _normalize(row.get("title"))
        section = str(row.get("section_key") or "")
        page_number = int(row.get("source_page") or 0)
        page = pages.get(page_number) or {}
        page_text_norm = _normalize(page.get("text"))

        # ``palco`` também aparece em frases como "presença de palco" e em
        # "Ficha de Palco". Esses casos não representam uma entidade cenográfica.
        if title_norm == "palco" and page_text_norm:
            structural_stage = any(signal in page_text_norm for signal in (
                "palco h", "palco com cenografia", "palco grande", "cenografia geral",
                "implantacao", "led 14x4", "led 6x2", "housemix",
            ))
            artistic_page = any(signal in page_text_norm for signal in (
                "mestre de cerimonias", "palestrante", "comunicadora", "comunicador",
                "apresentadora", "apresentador",
            ))
            if "ficha de palco" in page_text_norm and not structural_stage:
                row["title"] = "Ficha de palco"
                row["section_key"] = "communication"
                row["item_type"] = "Material impresso / comunicação"
                title_norm = "ficha de palco"
                section = "communication"
            elif artistic_page and not structural_stage:
                continue

        # Fichas genéricas de cobertura não são conteúdo de projeto.
        if generic_strategy.match(title_norm) or (
            str(row.get("extraction_origin") or "") == "automatic_repair"
            and title_norm.startswith("estrategia slide")
        ):
            continue

        # O título pesa primeiro. Isso impede que um item "Palco" seja contaminado
        # pelo texto completo de uma planta que também contém PHOTO OP, buffet etc.
        if title_norm == "ficha de palco":
            row["section_key"] = "communication"
            row["item_type"] = "Material impresso / comunicação"
        elif "photo op" in title_norm or "ponto de foto" in title_norm:
            row["section_key"] = "activations"
            row["item_type"] = "Photo-op"
        elif any(term in title_norm for term in ("quick massage", "ilha de massagem", "bacio di latte", "bar de cafe")):
            row["section_key"] = "activations"
            row["item_type"] = "Ativação / experiência"
        elif any(term in title_norm for term in ("jantar tematico", "banda", "studio 4", "musica ao vivo", "palestrante", "mestre de cerimonias")):
            row["section_key"] = "content_agenda"
            row["item_type"] = "Conteúdo artístico" if any(term in title_norm for term in artistic_terms) else "Conteúdo / programação"
        elif any(term in title_norm for term in ("kit de boas vindas", "camiseta", "mala", "mochila", "brinde", "gift", "press kit", "presskit")):
            row["section_key"] = "gifts"
            row["item_type"] = "Brinde / kit"
        elif any(term in title_norm for term in ("palco", "totem", "cenografia", "implantacao", "backdrop")):
            row["section_key"] = "scenography"
            row["item_type"] = "Cenografia / ambiente"
        elif any(term in title_norm for term in ("bourbon atibaia", "hotel bourbon", "ballroom garden", "sala londrina", "board room")):
            row["section_key"] = "journey_operation"
            row["item_type"] = "Local / operação"
        elif any(term in text for term in food_terms) or "jantar tematico" in text:
            row["section_key"] = "content_agenda"
            row["item_type"] = "Conteúdo / programação"
        elif any(term in text for term in artistic_terms):
            row["section_key"] = "content_agenda"
            row["item_type"] = "Conteúdo artístico"
        elif any(term in text for term in activation_terms):
            row["section_key"] = "activations"
            row["item_type"] = "Ativação / experiência"
        elif any(term in text for term in gift_terms):
            row["section_key"] = "gifts"
            row["item_type"] = "Brinde / kit"
        elif any(term in text for term in scenography_terms):
            row["section_key"] = "scenography"
            row["item_type"] = "Cenografia / ambiente"
        elif any(term in text for term in venue_terms):
            row["section_key"] = "journey_operation"
            row["item_type"] = "Local / operação"
        elif section == "gifts" and any(term in text for term in schedule_terms):
            row["section_key"] = "content_agenda"
            row["item_type"] = "Conteúdo / programação"
        elif section == "strategy" and any(term in text for term in schedule_terms):
            row["section_key"] = "content_agenda"
            row["item_type"] = "Conteúdo / programação"
        elif section == "strategy" and not any(term in text for term in strategy_terms):
            # Não removemos conteúdo desconhecido de forma agressiva; apenas evitamos
            # que contexto operacional evidente seja rotulado como estratégia.
            if "hotel" in text or "montagem" in text or "sala " in text:
                row["section_key"] = "journey_operation"
                row["item_type"] = "Local / operação"

        # Uma apresentação de proposta não comprova aprovação nem execução.
        if str(row.get("status") or "Não identificado") not in {"Referência", "Opção", "Recomendado"}:
            row["status"] = "Proposto"

        pattern = _semantic_pattern(title_norm) or _semantic_pattern(text)
        if pattern and page.get("text"):
            scoped = _scoped_page_text(str(page.get("text") or ""), pattern)
            if scoped:
                row["summary"] = scoped
                if len(str(row.get("description") or "")) > 520 or not row.get("description"):
                    row["description"] = scoped
                row["evidence"] = _clip(scoped, 260)

        cleaned.append(row)
    return cleaned


def _materialize_presentation_semantic(client: Any, source_file: Mapping[str, Any], data: bytes, warnings: list[str]) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    existing = _table_select_one(client, "memory_documents", project_id=project_id, content_sha256=sha)
    if existing:
        return {"memory_documents": 0, "memory_items": 0}
    doc = _source_document(source_file, data)
    if doc.mime_type != "application/pdf":
        warnings.append("A leitura visual especializada exige PDF; foi aplicada leitura textual de segurança.")
        text = doc.data.decode("utf-8", errors="replace") if doc.mime_type.startswith("text/") else ""
        return _materialize_presentation(client, source_file, text, warnings)
    api_key, model = _ai_settings()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY não está configurada para a decupagem especializada da apresentação")
    batches = extract_memory([doc], api_key=api_key, model=model)
    extraction = merge_memory_batches(batches)
    items = _sanitize_memory_items(list(extraction.get("items") or []), list(extraction.get("page_inventory") or []))
    extraction["items"] = items
    result = save_memory_presentation(
        client,
        project_id=project_id,
        source_document=doc,
        extraction=extraction,
        selected_items=items,
        document_title=str(extraction.get("document_title") or source_file.get("file_name") or "Apresentação"),
        version_label="Importação de projeto",
        document_status="sent_to_client",
    )
    document_id = result.get("document_id")
    if document_id and result.get("status") != "duplicate":
        try:
            current = _rows(client.table("memory_documents").select("raw_data").eq("id", document_id).limit(1).execute())
            raw = dict((current[0].get("raw_data") if current else {}) or {})
            raw.update({"materialized_by": WORKFLOW_VERSION, "source_file_id": source_file.get("id"), "semantic_pipeline": True})
            client.table("memory_documents").update({"raw_data": _safe_json(raw)}).eq("id", document_id).execute()
        except Exception as exc:
            warnings.append(f"Marcador da apresentação: {exc}")
    return {"memory_documents": 0 if result.get("status") == "duplicate" else 1, "memory_items": int(result.get("items_saved") or 0)}


def _parse_cost_with_ai(source_file: Mapping[str, Any], data: bytes) -> CostWorkbookResult:
    api_key, model = _ai_settings()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY não está configurada para a leitura semântica da planilha")
    doc = _source_document(source_file, data)
    ai_client = get_client(api_key)
    prompt = """
Você estrutura uma planilha humana de orçamento/verba de live marketing.
A planilha pode ter células mescladas, blocos, subtotais, categorias e textos livres.
Extraia itens reais de custo, não cabeçalhos. Regras:
- linhas como SUB TOTAL, TOTAL, EQUIPE E VERBAS ou nome de seção não são itens por si só;
- preserve a seção anterior em category;
- use números somente quando estiverem explícitos no arquivo;
- nunca converta ausência de valor em zero; use null;
- item_status: included, optional, reserve, pending, no_value ou client_responsibility;
- estimate_type: quoted, estimated, reserve, waiting_supplier ou no_value;
- client_total do documento deve ser null se nenhum valor monetário puder ser comprovado;
- não invente valores a partir de contexto externo.
Retorne CostWorkbookResult completo.
"""
    result = _structured_call(
        ai_client, model=model, prompt=prompt, docs=[doc],
        schema=CostWorkbookResult, context=f"planilha de custos {source_file.get('file_name')}"
    )
    result.file_name = str(source_file.get("file_name") or result.file_name)
    return result


def _materialize_costs_semantic(client: Any, source_file: Mapping[str, Any], data: bytes, warnings: list[str]) -> dict[str, int]:
    project_id = str(source_file.get("project_id") or "")
    sha = str(source_file.get("sha256") or "")
    existing = _table_select_one(client, "memory_cost_documents", project_id=project_id, content_sha256=sha)
    if existing:
        return {"cost_documents": 0, "cost_items": 0}
    file_name = str(source_file.get("file_name") or "planilha.xlsx")
    try:
        parsed = parse_cost_workbook(file_name, data)
        has_meaningful_structure = bool(parsed.items)
        obvious_garbage = all(
            _normalize(item.item_name) in {"sub total", "subtotal", "total", "equipe e verbas"}
            for item in parsed.items
        ) if parsed.items else True
        if not has_meaningful_structure or obvious_garbage:
            raise ValueError("a leitura tabular não encontrou itens reais")
    except Exception as deterministic_exc:
        warnings.append(f"A leitura tabular exigiu interpretação semântica: {deterministic_exc}")
        parsed = _parse_cost_with_ai(source_file, data)
    try:
        memory_rows = _rows(client.table("memory_items").select("*").eq("project_id", project_id).execute())
    except Exception:
        memory_rows = []
    result = save_cost_document(
        client,
        project_id=project_id,
        file_name=str(source_file.get("file_name") or "planilha.xlsx"),
        file_bytes=data,
        parsed=parsed,
        memory_items=memory_rows,
    )
    document_id = result.get("document_id")
    if document_id and result.get("status") != "duplicate":
        try:
            current = _rows(client.table("memory_cost_documents").select("metadata").eq("id", document_id).limit(1).execute())
            metadata = dict((current[0].get("metadata") if current else {}) or {})
            metadata.update({
                "materialized_by": WORKFLOW_VERSION,
                "source_file_id": source_file.get("id"),
                "semantic_pipeline": True,
                "document_role": source_file.get("document_role"),
            })
            client.table("memory_cost_documents").update({"metadata": _safe_json(metadata)}).eq("id", document_id).execute()
        except Exception as exc:
            warnings.append(f"Marcador da planilha: {exc}")
    warnings.extend(list(parsed.warnings or []))
    return {"cost_documents": 0 if result.get("status") == "duplicate" else 1, "cost_items": int(result.get("items_saved") or 0)}


def reprocess_project_semantically(client: Any, project_id: str) -> dict[str, Any]:
    try:
        files = _rows(
            client.table("source_files").select("*")
            .eq("project_id", project_id)
            .order("created_at", desc=False)
            .execute()
        )
    except Exception as exc:
        return {"project_id": project_id, "processed": 0, "errors": 1, "warnings": [str(exc)], "results": []}

    repair_warnings: list[str] = []
    repaired_files = [
        _repair_general_source_role(client, row, repair_warnings)
        for row in files
    ]
    repaired_files = _repair_bundle_topology(client, repaired_files, repair_warnings)
    relevant = [row for row in repaired_files if str(row.get("document_role") or "") in {
        "briefing_original", "detailed_costs", "preliminary_budget", "proposal_presentation", "final_presentation"
    }]
    results = []
    for row in relevant:
        _delete_legacy_specialized_rows(client, row)
        result = materialize_source_file(client, row, force_semantic_reprocess=True)
        results.append(result.as_dict())
    return {
        "project_id": project_id,
        "processed": sum(1 for r in results if r.get("status") != "error"),
        "errors": sum(1 for r in results if r.get("status") == "error"),
        "warnings": (repair_warnings + [w for r in results for w in (r.get("warnings") or [])])[:50],
        "results": results,
    }

def materialize_source_file(
    client: Any,
    source_file: Mapping[str, Any],
    *,
    source_bytes: bytes | None = None,
    text_override: str | None = None,
    force_semantic_reprocess: bool = False,
) -> MaterializationResult:
    source_file = dict(source_file)
    source_file_id = str(source_file.get("id") or "").strip()
    project_id = str(source_file.get("project_id") or "").strip()
    role = str(source_file.get("document_role") or "complementary_document").strip()
    if not source_file_id or not project_id:
        raise RuntimeError("source_file sem identidade de projeto suficiente para materialização")

    warnings: list[str] = []
    created: dict[str, int] = {}
    _sync_project_file(client, source_file_id, warnings)
    _ensure_project_file_role(client, source_file, warnings)
    text = _clean_text(text_override if text_override is not None else source_file.get("text_excerpt"))
    if source_bytes is None and role in {"briefing_original", "detailed_costs", "preliminary_budget", "proposal_presentation", "final_presentation"}:
        source_bytes = _download_bytes(client, source_file)

    try:
        if force_semantic_reprocess:
            _delete_legacy_specialized_rows(client, source_file)
        if role == "briefing_original":
            if source_bytes is None:
                raise RuntimeError("o briefing original não pôde ser recuperado do Storage")
            created.update(_materialize_briefing_semantic(client, source_file, source_bytes, warnings))
        elif role in {"detailed_costs", "preliminary_budget"}:
            if source_bytes is None:
                raise RuntimeError("a planilha original não pôde ser recuperada do Storage")
            created.update(_materialize_costs_semantic(client, source_file, source_bytes, warnings))
        elif role in {"proposal_presentation", "final_presentation"}:
            if source_bytes is None:
                raise RuntimeError("a apresentação original não pôde ser recuperada do Storage")
            created.update(_materialize_presentation_semantic(client, source_file, source_bytes, warnings))
        elif role == "feedback_approval":
            created.update(_materialize_feedback(client, source_file, text))
        elif role == "post_event_report":
            created.update(_materialize_report(client, source_file, text))
        # supplier_reference e complementary_document continuam visíveis em
        # Documentos via project_files/source_files, sem fabricar conteúdo.

        status = "materialized_with_warnings" if warnings else "materialized"
        notes = " | ".join(warnings[:12]) if warnings else f"Materializado pela NAVE {WORKFLOW_VERSION}."
        _mark_source_file(client, source_file_id, status=status, notes=notes, created=created)
        return MaterializationResult(source_file_id, project_id, role, status, created, warnings)
    except Exception as exc:
        warnings.append(str(exc))
        _mark_source_file(
            client,
            source_file_id,
            status="materialization_error",
            notes=" | ".join(warnings[:12]),
            created=created,
        )
        return MaterializationResult(source_file_id, project_id, role, "error", created, warnings)


def materialize_new_source_files(
    client: Any,
    source_files: Sequence[Mapping[str, Any]],
    documents_by_sha: Mapping[str, Any] | None = None,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    documents_by_sha = documents_by_sha or {}
    for row in source_files:
        sha = str(row.get("sha256") or "")
        document = documents_by_sha.get(sha)
        source_bytes = getattr(document, "data", None) if document is not None else None
        text = getattr(document, "text_excerpt", None) if document is not None else None
        result = materialize_source_file(
            client,
            row,
            source_bytes=source_bytes,
            text_override=text,
        )
        results.append(result.as_dict())
    return results


def repair_v2810_projects(client: Any, *, limit: int = MAX_SOURCE_FILES_REPAIR) -> dict[str, Any]:
    """Materializa arquivos importados pela V28.1.0 que ficaram só em source_files.

    É deliberadamente idempotente: as tabelas especializadas usam o hash do
    documento e a RPC de project_files também evita duplicação.
    """
    try:
        response = (
            client.table("source_files")
            .select("*")
            .not_.is_("project_id", "null")
            .not_.is_("document_role", "null")
            .order("created_at", desc=False)
            .limit(max(1, int(limit)))
            .execute()
        )
        rows = _rows(response)
    except Exception as exc:
        return {"scanned": 0, "repaired": 0, "errors": 1, "warnings": [str(exc)], "results": []}

    pending: list[dict[str, Any]] = []
    for row in rows:
        status = str(row.get("processing_status") or "").strip()
        metadata = row.get("metadata") if isinstance(row.get("metadata"), Mapping) else {}
        marker = str(metadata.get("materialized_by") or "").strip()
        if marker and status in {"materialized", "materialized_with_warnings"}:
            continue
        pending.append(row)

    results: list[dict[str, Any]] = []
    for row in pending:
        result = materialize_source_file(client, row)
        results.append(result.as_dict())

    repaired = sum(1 for item in results if item["status"] != "error")
    errors = sum(1 for item in results if item["status"] == "error")
    warnings = [warning for item in results for warning in item.get("warnings", [])]
    return {
        "scanned": len(rows),
        "pending": len(pending),
        "repaired": repaired,
        "errors": errors,
        "warnings": warnings[:40],
        "results": results,
    }
