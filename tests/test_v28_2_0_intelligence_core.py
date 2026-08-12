from __future__ import annotations

import io

from openpyxl import Workbook

from memory_cost_parser import parse_cost_workbook
from project_analyst import (
    FeedbackClaim,
    best_item_for_claim,
    derive_advanced_project_insights,
)
from project_batch_ingestion import classify_document


def test_jovi_like_briefing_is_not_confused_with_post_event_deliverable_terms():
    text = """
    JOB: X300 Series Launch Event
    INFOS GERAIS
    CONCORRENCIA: SIM
    BRIEFING
    PUBLICO ALVO: criadores de conteúdo e jornalistas
    OBJETIVO: lançamento nacional
    ENTREGÁVEIS: proposta criativa e logística
    OBRIGATORIEDADES: orçamento detalhado e relatório pós-evento com resultados e aprendizados
    FINANCEIRO / BUDGET: R$ 1.300.000,00
    """
    role, confidence, _ = classify_document("VOE_Briefing_JOVI_X300_V2.docx", text)
    assert role == "briefing_original"
    assert confidence >= 0.9


def test_english_proposal_pdf_is_not_confused_with_supplier_portfolio():
    text = """
    BRIEF RECAP | OUR GOAL
    OUR CHALLENGE
    BRAND CONTEXT
    COMPETITOR LANDSCAPE
    OUR APPROACH
    INSIGHT
    JOVI X300 SERIES ON TOUR
    EVENT JOURNEY
    PRESS KIT
    CREATOR RECOMMENDATION
    PRODUCT REVEAL
    YouTube activation, Instagram activation, TikTok activation, Kwai activation
    portfolio reference
    """
    role, confidence, _ = classify_document("PDF_LANCAMENTO_JOVI_X300_30.06.pdf", text)
    assert role == "proposal_presentation"
    assert confidence >= 0.9


def _english_cost_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "02_Cost_Breakdown"
    headers = [
        None, "No", "Category", "Item Description", "Service Description Scope",
        "Vendor/Subcontractor", "Limitations Notes", "Cost Classification", "Unit",
        "Quantity", "Day/Shift", "Price", "Unit Cost (Before tax)", "Agency Markup%",
        "Total (Before tax)", "Unit price (including tax, XX%)",
        "Total (including tax, XX%)",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(6, col, value)
    # subtotal: deve ser ignorado
    subtotal = [None, 1, "Venue & Infrastructure", None, None, None, None, None, "Subtotal", None, None, None, 1000, 100, 1100, 25, 1125]
    item = [None, "1.1", "Venue & Infrastructure", "Venue Rental", "Main hall", "Venue SA", None, "PVCH", None, 1, 1, 1000, 1000, 100, 1100, 25, 1125]
    item2 = [None, "2.1", "Event Production", "Scenography", "Scenic build", "Scenic Co", None, "PVCH", None, 1, 1, 2000, 2000, 200, 2200, 50, 2250]
    total = [None, "Total Cost (Without Tax)", None, None, None, None, None, None, None, None, None, None, 3000, 300, 3300, 75, 3375]
    for r, values in enumerate((subtotal, item, item2, total), start=7):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_english_jovi_style_cost_template_is_structured_and_reconciled():
    parsed = parse_cost_workbook("JOVI_Quotation.xlsx", _english_cost_workbook())
    assert parsed.sheet_name == "02_Cost_Breakdown"
    assert len(parsed.items) == 2
    assert parsed.client_total == 3375
    assert parsed.total_base == 3000
    assert parsed.fees_total == 300
    assert parsed.charges_total == 75
    assert parsed.items[0].category == "Venue & Infrastructure"
    assert parsed.items[0].item_name == "Venue Rental"
    assert parsed.items[0].raw_data["vendor"] == "Venue SA"


def test_feedback_claim_links_to_named_solution_not_generic_neighbor():
    claim = FeedbackClaim(
        title="Concept & Campaign Alignment",
        theme="creative_concept",
        sentiment="positive",
        evidence_quote="We highly commend your concept JOVI X300 Series On Tour",
        related_entities=["JOVI X300 Series On Tour"],
        item_outcome_status="approved",
    )
    items = [
        {"id": "1", "title": "JOVI X300 Series On Tour", "summary": "Multiple lenses to capture Brazilian stories", "section_key": "strategy"},
        {"id": "2", "title": "Cinemateca", "summary": "Venue proposal", "section_key": "journey_operation"},
    ]
    item, score = best_item_for_claim(claim, items)
    assert item and item["id"] == "1"
    assert score >= 0.5


def test_project_analyst_crosses_budget_cost_feedback_and_briefing():
    snapshot = {
        "briefing_documents": [{"metadata": {"audience_quantity": "250 convidados"}, "budget_amount": 1_300_000}],
        "briefing_requirements": [{"id": "r1", "title": "Local para 250 convidados", "mandatory": True, "priority": "critical"}],
        "memory_items": [{"id": "m1", "title": "Cinemateca", "section_key": "journey_operation"}],
        "cost_items": [
            {"id": "c1", "category": "Venue & Infrastructure", "item_name": "Venue Rental", "client_total": 240_000},
            {"id": "c2", "category": "Event Production", "item_name": "Scenography", "client_total": 420_000},
            {"id": "c3", "category": "AV & Technical", "item_name": "LED", "client_total": 230_000},
            {"id": "c4", "category": "Staffing & Talent", "item_name": "Staff", "client_total": 210_000},
            {"id": "c5", "category": "Other", "item_name": "Other", "client_total": 400_000},
        ],
        "cost_links": [{"memory_item_id": "m1", "cost_item_id": "c1", "link_status": "confirmed"}],
        "item_outcomes": [{"item_id": "m1", "outcome_status": "not_approved", "feedback_summary": "Capacity constraints"}],
        "briefing_links": [{"requirement_id": "r1", "memory_item_id": "m1", "link_status": "confirmed", "adherence_status": "not_fulfilled"}],
        "feedback_entries": [{"theme": "operation", "sentiment": "negative"}],
    }
    result = derive_advanced_project_insights(snapshot, proposal_total=1_500_000, budget_amount=1_300_000)
    assert result["audience_quantity"] == 250
    assert result["cost_per_attendee"] == 6000
    assert result["challenged_items"][0]["title"] == "Cinemateca"
    assert result["requirement_risks"]
    assert any(row["title"] == "Crítica em investimento relevante" for row in result["findings"])
    assert any(row["title"] == "Aderência financeira" for row in result["findings"])
