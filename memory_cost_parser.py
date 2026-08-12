from __future__ import annotations

import csv
import io
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from memory_learning_models import (
    CostItem,
    CostWorkbookResult,
)


HEADER_ALIASES = {
    "code": {"#", "codigo", "cod", "item codigo", "no", "number"},
    # V28.2 distingue categoria e nome do item. Em planilhas antigas ambos
    # continuam aceitos pelo campo legado category_or_name.
    "category": {
        "categoria", "grupo", "secao", "centro de custo", "category",
    },
    "item_name": {
        "item description", "item name", "nome do item", "descricao do item",
        "produto servico", "service item",
    },
    "category_or_name": {
        "item descricao", "item", "categoria grupo", "grupo item",
    },
    "description": {
        "descricao", "escopo", "especificacao", "detalhamento",
        "service description scope", "service description", "scope",
        "description",
    },
    "vendor": {
        "vendor subcontractor", "vendor", "subcontractor", "fornecedor",
        "fornecedor subcontratado",
    },
    "limitations": {
        "limitations notes", "limitations", "notes", "observacoes",
        "observacao",
    },
    "billing_type": {
        "tipo faturamento", "faturamento", "tipo de faturamento",
        "cost classification", "classification",
    },
    "unit": {"unit", "unidade", "un"},
    "quantity": {"quant", "qtd", "qtde", "quantidade", "quantity", "qty"},
    "period": {
        "periodo", "diarias", "dias", "duracao", "day shift", "days shift",
        "day", "shift",
    },
    "unit_value": {
        "valor unit", "valor unitario", "unitario", "preco unitario",
        "price", "unit price",
    },
    "base_value": {
        "valor total", "custo base", "subtotal",
        "unit cost before tax", "cost before tax",
    },
    "fees_value": {
        "honorarios", "honorario", "fee", "agency markup", "agency markup percent",
        "agency markup %",
    },
    # Alguns templates internacionais usam o cabeçalho enganoso
    # "Unit price (including tax, XX%)" para a PARCELA DE IMPOSTOS.
    "charges_value": {
        "encargos", "impostos", "taxas", "tax", "tax amount",
        "unit price including tax xx", "unit price including tax xx percent",
    },
    "pre_tax_total": {
        "total before tax", "total sem imposto", "total pre tax",
    },
    "client_total": {
        "total com honorarios e encargos", "total cliente", "valor final",
        "total final", "total including tax xx", "total including tax xx percent",
        "total including tax", "grand total",
    },
}


STATUS_KEYWORDS = {
    "client_responsibility": [
        "responsabilidade cliente",
        "responsabilidade do cliente",
    ],
    "optional": ["opcional"],
    "reserve": [
        "reserva de verba",
        "verba de reserva",
    ],
    "pending": [
        "aguardando",
        "a definir",
        "em definicao",
        "apos visita tecnica",
        "a avaliar",
    ],
}

ESTIMATE_KEYWORDS = {
    "estimated": [
        "estimado",
        "estimativa",
        "pode sofrer alteracao",
        "sujeito a alteracao",
    ],
    "waiting_supplier": [
        "aguardando fornecedor",
        "aguardando lista",
        "fornecedor oficial",
    ],
}


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize(
        "NFKD",
        str(value or ""),
    )
    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
    )
    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text.casefold(),
    )
    return re.sub(r"\s+", " ", text).strip()


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = re.sub(
        r"[^\d,.\-]",
        "",
        str(value).strip(),
    )

    if not text:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def _first_line(value: Any) -> str:
    lines = [
        re.sub(r"\s+", " ", line).strip(" -*•\t")
        for line in str(value or "").splitlines()
        if line.strip()
    ]

    if not lines:
        return ""

    result = re.sub(
        r"^(gift|ativacao|ativação)\s*[-:]\s*",
        "",
        lines[0],
        flags=re.IGNORECASE,
    )
    return result[:220]


def _sheet_matrix_xlsx(
    data: bytes,
    *,
    keep_vba: bool,
) -> list[tuple[str, list[list[Any]]]]:
    workbook = load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=True,
        keep_vba=keep_vba,
    )
    sheets = []

    try:
        for worksheet in workbook.worksheets:
            matrix = []

            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row or 1, 5000),
                values_only=True,
            ):
                matrix.append(list(row[:60]))

            sheets.append((worksheet.title, matrix))
    finally:
        workbook.close()

    return sheets


def _sheet_matrix_xls(
    data: bytes,
) -> list[tuple[str, list[list[Any]]]]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=data)
    return [
        (
            sheet.name,
            [
                sheet.row_values(row_index)
                for row_index in range(min(sheet.nrows, 5000))
            ],
        )
        for sheet in workbook.sheets()
    ]


def _sheet_matrix_csv(
    data: bytes,
) -> list[tuple[str, list[list[Any]]]]:
    decoded = data.decode("utf-8-sig", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(
            decoded[:5000],
            delimiters=",;\t|",
        )
    except csv.Error:
        dialect = csv.excel

    return [
        (
            "Planilha",
            list(csv.reader(io.StringIO(decoded), dialect)),
        )
    ]


def _load_sheets(
    file_name: str,
    data: bytes,
) -> list[tuple[str, list[list[Any]]]]:
    suffix = Path(file_name).suffix.casefold()

    if suffix in {".xlsx", ".xlsm"}:
        return _sheet_matrix_xlsx(
            data,
            keep_vba=suffix == ".xlsm",
        )

    if suffix == ".xls":
        return _sheet_matrix_xls(data)

    if suffix == ".csv":
        return _sheet_matrix_csv(data)

    raise ValueError(
        "Formato de planilha não suportado. "
        "Use XLSX, XLSM, XLS ou CSV."
    )


def _header_key(value: Any) -> str | None:
    raw = str(value or "").strip()
    if raw == "#":
        return "code"

    normalized = normalize_text(value)

    for key, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return key

    return None


def _find_header(
    matrix: list[list[Any]],
) -> tuple[int, dict[str, int]]:
    best_row = -1
    best_map: dict[str, int] = {}

    for row_index, row in enumerate(matrix[:120], start=1):
        mapping: dict[str, int] = {}

        for column_index, value in enumerate(row[:40]):
            key = _header_key(value)

            if key and key not in mapping:
                mapping[key] = column_index

        required_score = sum(
            key in mapping
            for key in {
                "description",
                "item_name",
                "quantity",
                "unit_value",
                "base_value",
                "client_total",
                "category",
                "category_or_name",
            }
        )

        if (
            len(mapping) >= 4
            and required_score >= 2
            and len(mapping) > len(best_map)
        ):
            best_row = row_index
            best_map = mapping

    if best_row < 1:
        raise ValueError(
            "Não foi possível identificar uma "
            "tabela de custos na planilha."
        )

    return best_row, best_map


def _sheet_score(matrix: list[list[Any]]) -> int:
    try:
        _, mapping = _find_header(matrix)
    except ValueError:
        return 0

    return len(mapping) * 10 + min(len(matrix), 200)


def _get(
    row: list[Any],
    mapping: dict[str, int],
    key: str,
) -> Any:
    index = mapping.get(key)

    if index is None or index >= len(row):
        return None

    return row[index]


def _metadata_before_header(
    matrix: list[list[Any]],
    header_row: int,
) -> dict:
    metadata: dict[str, Any] = {}
    title_candidates = []

    for row in matrix[: max(0, header_row - 1)]:
        values = [
            str(value).strip()
            for value in row
            if value not in (None, "")
        ]

        for value in values:
            if ":" in value:
                key, content = value.split(":", 1)

                if content.strip():
                    metadata[normalize_text(key)] = content.strip()
            elif (
                len(value) >= 12
                and not re.fullmatch(r"[\d\s./%-]+", value)
            ):
                title_candidates.append(value)

    metadata["_title_candidates"] = title_candidates
    return metadata


def _category_label(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
    label = lines[0] if lines else re.sub(r"\s+", " ", text).strip()
    return label[:120]


def _status_and_estimate(
    description: str,
    *,
    base_value: float | None,
    client_total: float | None,
) -> tuple[str, str, list[str]]:
    normalized = normalize_text(description)
    flags = []
    status = None

    for status_key, keywords in STATUS_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            status = status_key
            flags.append(status_key)
            break

    if status is None:
        status = (
            "included"
            if (client_total or 0) > 0 or (base_value or 0) > 0
            else "no_value"
        )

    estimate_type = None

    for estimate_key, keywords in ESTIMATE_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            estimate_type = estimate_key
            flags.append(estimate_key)
            break

    if estimate_type is None:
        if status == "reserve":
            estimate_type = "reserve"
        elif (client_total or 0) > 0 or (base_value or 0) > 0:
            estimate_type = "quoted"
        else:
            estimate_type = "no_value"

    return status, estimate_type, list(dict.fromkeys(flags))


def _extract_total_row(
    matrix: list[list[Any]],
    header_row: int,
    mapping: dict[str, int],
) -> dict:
    # O total real costuma estar no fim da planilha, não nas oito primeiras
    # linhas após o cabeçalho. Priorizamos TOTAL GERAL / TOTAL FINAL e, na
    # ausência deles, o último TOTAL com valor monetário explícito.
    candidates: list[tuple[int, int, dict[str, float | None]]] = []
    for row_index, row in enumerate(matrix[header_row:], start=header_row + 1):
        first_values = " ".join(normalize_text(value) for value in row[:6]).strip()
        if not re.search(r"\btotal\b", first_values):
            continue
        values = {
            "total_base": _number(_get(row, mapping, "base_value")),
            "fees_total": _number(_get(row, mapping, "fees_value")),
            "charges_total": _number(_get(row, mapping, "charges_value")),
            "client_total": _number(_get(row, mapping, "client_total")),
        }
        if not any(value is not None for value in values.values()):
            continue
        priority = 3 if any(term in first_values for term in ("total geral", "total final", "grand total")) else 2
        if "sub total" in first_values or "subtotal" in first_values:
            priority = 1
        candidates.append((priority, row_index, values))

    if not candidates:
        return {}
    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[-1][2]




def _parse_preliminary_budget_matrix(
    file_name: str,
    sheets: list[tuple[str, list[list[Any]]]],
) -> CostWorkbookResult | None:
    """Lê estudos humanos de verba sem cabeçalho tabular convencional.

    O padrão observado no Planeja 27 usa rótulo, valor e percentual espalhados
    em colunas visuais. A ausência de uma linha de cabeçalho não torna esses
    números inconclusivos: a NAVE deve reconstruir a estrutura com segurança.
    """
    best: tuple[str, list[tuple[int, str, float | None, float | None]]] | None = None
    for sheet_name, matrix in sheets:
        rows: list[tuple[int, str, float | None, float | None]] = []
        for row_number, row in enumerate(matrix[:160], start=1):
            populated = [(idx, value) for idx, value in enumerate(row[:12]) if value not in (None, "") and str(value).strip()]
            if not populated:
                continue
            text_cells = [(idx, str(value).strip()) for idx, value in populated if not isinstance(value, (int, float))]
            if not text_cells:
                continue
            label_index, label = text_cells[0]
            normalized = normalize_text(label)
            if not normalized:
                continue
            numeric_after = [
                _number(value)
                for idx, value in populated
                if idx > label_index and _number(value) is not None
            ]
            amount = numeric_after[0] if numeric_after else None
            pct = numeric_after[1] if len(numeric_after) > 1 else None
            if pct is not None and abs(pct) > 1.5:
                pct = None
            rows.append((row_number, label, amount, pct))

        normalized_labels = [normalize_text(row[1]) for row in rows]
        budget_rows = [row for row in rows if normalize_text(row[1]) == "budget" and row[2] is not None]
        meaningful = [
            row for row in rows
            if normalize_text(row[1]) not in {"budget", "sub total", "subtotal"}
            and row[2] is not None
        ]
        if budget_rows and len(meaningful) >= 4 and any(
            signal in normalized_labels
            for signal in ("fee", "impostos", "cenografia e infraestrutura", "tecnica", "atracoes")
        ):
            if best is None or len(rows) > len(best[1]):
                best = (sheet_name, rows)

    if best is None:
        return None

    sheet_name, rows = best
    budget_values = [float(amount) for _, label, amount, _ in rows if normalize_text(label) == "budget" and amount is not None]
    budget_amount = max(budget_values) if budget_values else None
    items: list[CostItem] = []
    percentage_total = 0.0
    allocation_total = 0.0
    for row_number, label, amount, pct in rows:
        normalized = normalize_text(label)
        if normalized in {"budget", "sub total", "subtotal"}:
            continue
        if amount is None:
            continue
        percentage_total += float(pct or 0)
        allocation_total += float(amount or 0)
        items.append(
            CostItem(
                source_sheet=sheet_name,
                source_row=row_number,
                item_code=None,
                category="Estudo de verba",
                item_name=label.strip(),
                description=label.strip(),
                quantity=None,
                period=None,
                unit_value=None,
                base_value=float(amount),
                fees_value=None,
                charges_value=None,
                client_total=float(amount),
                item_status="included",
                estimate_type="estimated",
                flags=["preliminary_budget"],
                raw_data={"allocation_pct": pct, "explicit_zero": float(amount) == 0.0},
            )
        )

    if not items:
        return None
    reconciled_total = allocation_total
    client_total = budget_amount if budget_amount is not None else reconciled_total
    warnings: list[str] = []
    if budget_amount is not None and abs(reconciled_total - budget_amount) > max(1.0, budget_amount * 0.01):
        warnings.append(
            "O estudo de verba foi lido como matriz de alocação, mas a soma das categorias não reconciliou exatamente com o budget informado."
        )

    return CostWorkbookResult(
        file_name=file_name,
        title=Path(file_name).stem,
        sheet_name=sheet_name,
        header_row=1,
        project_name=Path(file_name).stem,
        total_base=reconciled_total,
        client_total=client_total,
        items=items,
        warnings=warnings,
        metadata={
            "cost_kind": "preliminary_budget",
            "budget_amount": budget_amount,
            "allocation_total": reconciled_total,
            "allocation_percentage_total": percentage_total,
            "parser_mode": "visual_allocation_matrix",
        },
    )


def _additional_sheet_totals(
    sheets: list[tuple[str, list[list[Any]]]],
    selected_sheet_name: str,
) -> list[dict[str, Any]]:
    totals: list[dict[str, Any]] = []
    for sheet_name, matrix in sheets:
        if sheet_name == selected_sheet_name or _sheet_score(matrix) <= 0:
            continue
        try:
            header_row, mapping = _find_header(matrix)
            values = _extract_total_row(matrix, header_row, mapping)
        except Exception:
            continue
        client_total = values.get("client_total")
        if client_total is None:
            continue
        totals.append({
            "sheet_name": sheet_name,
            "client_total": float(client_total),
            "total_base": values.get("total_base"),
            "fees_total": values.get("fees_total"),
            "charges_total": values.get("charges_total"),
        })
    return totals


def parse_cost_workbook(
    file_name: str,
    data: bytes,
) -> CostWorkbookResult:
    sheets = _load_sheets(file_name, data)

    if not sheets:
        raise ValueError("A planilha não possui abas legíveis.")

    if max((_sheet_score(matrix) for _, matrix in sheets), default=0) <= 0:
        preliminary = _parse_preliminary_budget_matrix(file_name, sheets)
        if preliminary is not None:
            return preliminary

    selected_sheet_name, matrix = max(
        sheets,
        key=lambda item: _sheet_score(item[1]),
    )
    header_row, mapping = _find_header(matrix)
    metadata = _metadata_before_header(matrix, header_row)
    totals = _extract_total_row(matrix, header_row, mapping)

    header_values = (
        matrix[header_row - 1]
        if header_row - 1 < len(matrix)
        else []
    )
    mapped_columns = set(mapping.values())
    unknown_columns = [
        str(value).strip()
        for index, value in enumerate(header_values)
        if (
            value not in (None, "")
            and index not in mapped_columns
        )
    ]

    items = []
    current_category = None
    blank_streak = 0

    for source_row, row in enumerate(
        matrix[header_row:],
        start=header_row + 1,
    ):
        row_text = " ".join(
            str(value or "").strip()
            for value in row
        ).strip()

        if not row_text:
            blank_streak += 1

            if blank_streak >= 5 and items:
                break
            continue

        blank_streak = 0

        code = _get(row, mapping, "code")
        category = _get(row, mapping, "category")
        item_name_value = _get(row, mapping, "item_name")
        category_or_name = _get(row, mapping, "category_or_name")
        description = _get(row, mapping, "description")
        vendor = _get(row, mapping, "vendor")
        limitations = _get(row, mapping, "limitations")
        billing_type = _get(row, mapping, "billing_type")
        unit = _get(row, mapping, "unit")
        quantity = _number(_get(row, mapping, "quantity"))
        period = _number(_get(row, mapping, "period"))
        unit_value = _number(_get(row, mapping, "unit_value"))
        base_value = _number(_get(row, mapping, "base_value"))
        fees_value = _number(_get(row, mapping, "fees_value"))
        charges_value = _number(_get(row, mapping, "charges_value"))
        pre_tax_total = _number(_get(row, mapping, "pre_tax_total"))
        client_total = _number(_get(row, mapping, "client_total"))

        normalized_first = normalize_text(
            " ".join(str(value or "") for value in row[:4])
        )

        if re.search(r"\btotal\b", normalized_first):
            continue

        # Templates internacionais usam linhas de subtotal com categoria +
        # valores, mas sem Item Description. Elas não são entregas/custos
        # independentes e seriam duplicadas com as linhas detalhadas.
        if normalize_text(unit) in {"subtotal", "sub total"} and not str(item_name_value or "").strip():
            current_category = _category_label(category) or current_category
            continue

        nonempty_values = [
            value for value in row
            if value not in (None, "") and str(value).strip()
        ]
        numeric_values = [
            quantity, period, unit_value, base_value,
            fees_value, charges_value, pre_tax_total, client_total,
        ]
        only_text = str(nonempty_values[0]).strip() if len(nonempty_values) == 1 else ""
        letters = "".join(ch for ch in only_text if ch.isalpha())
        looks_like_heading = bool(letters) and (letters.upper() == letters or only_text.endswith(":"))
        single_text_section = (
            len(nonempty_values) == 1
            and looks_like_heading
            and all(value in (None, 0) for value in numeric_values)
        )
        category_text = str(category or category_or_name or "").strip()
        description_text = str(description or "").strip()
        same_category_description = (
            bool(category_text)
            and normalize_text(category_text) == normalize_text(description_text)
            and not str(billing_type or "").strip()
            and all(value in (None, 0) for value in numeric_values)
        )
        is_category = (
            single_text_section
            or same_category_description
            or (
                bool(category_text)
                and not description_text
                and not str(billing_type or "").strip()
                and all(value in (None, 0) for value in numeric_values)
            )
        )

        if is_category:
            current_category = _category_label(category_text or description_text) or current_category
            continue

        full_description = str(
            description
            or item_name_value
            or category_or_name
            or ""
        ).strip()

        explicit_item_name = _first_line(item_name_value) if item_name_value not in (None, "") else ""
        item_name = explicit_item_name or _first_line(full_description)

        # Uma linha sem nome/descrição de entrega é metadado, subtotal ou
        # cabeçalho e não deve virar item financeiro.
        if not item_name:
            continue

        # Algumas planilhas usam apenas "Valor total" e deixam a coluna
        # "Total cliente" vazia. Quando os componentes monetários estão explícitos,
        # reconciliamos o total por item sem inventar nenhum valor externo.
        effective_client_total = client_total
        if effective_client_total is None and base_value is not None:
            effective_client_total = float(base_value)
            if fees_value is not None:
                effective_client_total += float(fees_value)
            if charges_value is not None:
                effective_client_total += float(charges_value)

        status, estimate_type, flags = _status_and_estimate(
            full_description,
            base_value=base_value,
            client_total=effective_client_total,
        )

        items.append(
            CostItem(
                source_sheet=selected_sheet_name,
                source_row=source_row,
                item_code=(
                    str(code).strip()
                    if code not in (None, "")
                    else None
                ),
                category=(
                    _category_label(category)
                    or current_category
                    or (
                        str(category_or_name).strip()
                        if category_or_name and description
                        else None
                    )
                ),
                item_name=item_name,
                description=full_description,
                billing_type=(
                    str(billing_type).strip()
                    if billing_type not in (None, "")
                    else None
                ),
                quantity=quantity,
                period=period,
                unit_value=unit_value,
                base_value=base_value,
                fees_value=fees_value,
                charges_value=charges_value,
                client_total=effective_client_total,
                item_status=status,
                estimate_type=estimate_type,
                flags=flags,
                raw_data={
                    "row": list(row),
                    "vendor": str(vendor).strip() if vendor not in (None, "") else None,
                    "limitations": str(limitations).strip() if limitations not in (None, "") else None,
                    "unit": str(unit).strip() if unit not in (None, "") else None,
                    "pre_tax_total": pre_tax_total,
                },
            )
        )

    if not items:
        raise ValueError(
            "A tabela foi localizada, mas nenhum "
            "item de custo pôde ser estruturado."
        )

    title_candidates = metadata.pop("_title_candidates", [])
    project_name = (
        title_candidates[0]
        if title_candidates
        else Path(file_name).stem
    )

    warnings = []

    if Path(file_name).suffix.casefold() == ".xlsm":
        warnings.append(
            "O arquivo contém suporte a macros. "
            "A NAVE leu somente valores e células; "
            "nenhuma macro foi executada."
        )

    if unknown_columns:
        warnings.append(
            "Algumas colunas não fazem parte da "
            "estrutura atual e foram preservadas "
            "apenas no arquivo original."
        )

    has_any_monetary_value = any(
        value is not None
        for item in items
        for value in (
            item.unit_value, item.base_value, item.fees_value,
            item.charges_value, item.client_total,
        )
    )
    calculated_total = (
        sum(float(item.client_total or 0) for item in items)
        if has_any_monetary_value
        else None
    )
    stated_total = totals.get("client_total")
    metadata["cost_kind"] = "detailed_costs"
    metadata["parser_mode"] = "structured_cost_table"
    additional_totals = _additional_sheet_totals(sheets, selected_sheet_name)
    if additional_totals:
        metadata["additional_sheet_totals"] = additional_totals

    if (
        stated_total is not None
        and calculated_total is not None
        and abs(calculated_total - stated_total)
        > max(1.0, stated_total * 0.01)
    ):
        warnings.append(
            "O total das linhas estruturadas não "
            "reconciliou exatamente com o total "
            "informado na planilha."
        )

    return CostWorkbookResult(
        file_name=file_name,
        title=Path(file_name).stem,
        sheet_name=selected_sheet_name,
        header_row=header_row,
        project_name=project_name,
        event_date=metadata.get("data do evento"),
        presentation_date=metadata.get("data de apresentacao"),
        macros_present=(
            Path(file_name).suffix.casefold() == ".xlsm"
        ),
        total_base=totals.get("total_base"),
        fees_total=totals.get("fees_total"),
        charges_total=totals.get("charges_total"),
        client_total=(
            stated_total
            if stated_total is not None
            else calculated_total
        ),
        items=items,
        warnings=warnings,
        unknown_columns=unknown_columns,
        metadata=metadata,
    )
